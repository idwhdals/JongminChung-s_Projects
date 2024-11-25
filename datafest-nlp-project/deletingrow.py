import openpyxl

# Load the input workbook
workbook = openpyxl.load_workbook('qq.xlsx')

# Get the active worksheet
worksheet = workbook.active

# Get the maximum number of rows in the worksheet
max_row = worksheet.max_row

# Create a dictionary to store the grades for each ID
grades = {}

# Loop through all rows in the input worksheet, starting from row 2
for row in range(2, max_row + 1):
    cell_value = worksheet.cell(row=row, column=1).value
    if cell_value:
        cell_value = str(cell_value).lower()
        id_value = worksheet.cell(row=row, column=2).value
        if "god bless" in cell_value:
            if id_value in grades:
                grades[id_value] += 5
            else:
                grades[id_value] = 5
        if "wow" in cell_value:
            if id_value in grades:
                grades[id_value] += 3.5
            else:
                grades[id_value] = 3.5
        if "perfect" in cell_value or "confidence" in cell_value:
            if id_value in grades:
                grades[id_value] += 3
            else:
                grades[id_value] = 3
        if "oh" in cell_value or "happy" in cell_value:
            if id_value in grades:
                grades[id_value] += 2.5
            else:
                grades[id_value] = 2.5
        if "understand" in cell_value or "thank you for" in cell_value or "glad" in cell_value or "able to" in cell_value or "solve" in cell_value:
            if id_value in grades:
                grades[id_value] += 1.5
            else:
                grades[id_value] = 1.5
        if "unfortunately" in cell_value or "difficult" in cell_value or "unable" in cell_value:
            if id_value in grades:
                grades[id_value] -= 1
            else:
                grades[id_value] = -1
        if "worried" in cell_value or "hard" in cell_value:
            if id_value in grades:
                grades[id_value] -= 0.5
            else:
                grades[id_value] = -0.5
        if "sorry" in cell_value or "cannot" in cell_value or "late" in cell_value:
            if id_value in grades:
                grades[id_value] -= 0.25
            else:
                grades[id_value] = -0.25
 
# Create a new Excel workbook and worksheet
new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active
     
# Write the header row to the new worksheet
new_worksheet.append(['id', 'grade'])

# Loop through the grades dictionary and write each ID and grade to the new worksheet
for id_value, grade_value in grades.items():
    if grade_value != 0: # skip rows with zero grade
        new_worksheet.append([id_value, grade_value])

# Save the new workbook to a file
new_workbook.save('please.xlsx')